from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q, Count
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from .models import Employee, Customer
from farmers.models import Farmer
from customers.models import Customer as CustomerModel
import csv
import openpyxl
from datetime import datetime, timedelta


class ManagementIndexView(LoginRequiredMixin, TemplateView):
    """Trang chủ quản lý"""
    template_name = 'management/index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Thống kê
        context['employee_count'] = Employee.objects.filter(is_active=True).count()
        context['farmer_count'] = Farmer.objects.filter(is_active=True).count()
        context['customer_count'] = CustomerModel.objects.filter(is_active=True).count()
        
        return context


# Employee Views
class EmployeeListView(LoginRequiredMixin, ListView):
    """Danh sách nhân viên"""
    model = Employee
    template_name = 'management/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Employee.objects.all()
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(employee_id__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Filters
        department = self.request.GET.get('department')
        if department:
            queryset = queryset.filter(department=department)
            
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
            
        position = self.request.GET.get('position')
        if position:
            queryset = queryset.filter(position=position)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics
        context['total_employees'] = Employee.objects.count()
        context['active_employees'] = Employee.objects.filter(is_active=True).count()
        context['managers_count'] = Employee.objects.filter(position='manager').count()
        
        # New employees this month
        this_month = datetime.now().replace(day=1)
        context['new_employees'] = Employee.objects.filter(
            created_at__gte=this_month
        ).count()
        
        return context


class EmployeeCreateView(LoginRequiredMixin, CreateView):
    """Tạo nhân viên mới"""
    model = Employee
    template_name = 'management/employee_create.html'
    fields = [
        'employee_id', 'first_name', 'last_name', 'nickname', 'gender',
        'date_of_birth', 'national_id', 'avatar', 'email', 'phone', 'address',
        'emergency_contact_name', 'emergency_contact_phone', 'department',
        'position', 'hire_date', 'salary', 'job_description', 'education_level',
        'marital_status', 'bank_account', 'bank_name', 'is_active', 'notes'
    ]
    success_url = reverse_lazy('management:employee_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Tạo nhân viên mới thành công!')
        return super().form_valid(form)


class EmployeeDetailView(LoginRequiredMixin, DetailView):
    """Chi tiết nhân viên"""
    model = Employee
    template_name = 'management/employee_detail.html'
    context_object_name = 'employee'


class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    """Cập nhật nhân viên"""
    model = Employee
    template_name = 'management/employee_edit.html'
    fields = [
        'employee_id', 'first_name', 'last_name', 'nickname', 'gender',
        'date_of_birth', 'national_id', 'avatar', 'email', 'phone', 'address',
        'emergency_contact_name', 'emergency_contact_phone', 'department',
        'position', 'hire_date', 'salary', 'job_description', 'education_level',
        'marital_status', 'bank_account', 'bank_name', 'is_active', 'notes'
    ]
    success_url = reverse_lazy('management:employee_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Cập nhật nhân viên thành công!')
        return super().form_valid(form)


class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    """Xóa nhân viên"""
    model = Employee
    success_url = reverse_lazy('management:employee_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Xóa nhân viên thành công!')
        return super().delete(request, *args, **kwargs)


# Farmer Views
class FarmerListView(LoginRequiredMixin, ListView):
    """Danh sách trang trại"""
    model = Farmer
    template_name = 'management/farmer_list.html'
    context_object_name = 'farmers'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Farmer.objects.all()
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(farm_name__icontains=search) |
                Q(farmer_code__icontains=search) |
                Q(owner_name__icontains=search) |
                Q(email__icontains=search)
            )
        
        # Filters
        certification_type = self.request.GET.get('certification_type')
        if certification_type:
            queryset = queryset.filter(certification_type=certification_type)
            
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistics
        context['total_farmers'] = Farmer.objects.count()
        context['organic_farmers'] = Farmer.objects.filter(certification_type='organic').count()
        context['vietgap_farmers'] = Farmer.objects.filter(certification_type='vietgap').count()
        context['globalgap_farmers'] = Farmer.objects.filter(certification_type='globalgap').count()
        
        return context


class FarmerCreateView(LoginRequiredMixin, CreateView):
    """Tạo trang trại mới"""
    model = Farmer
    template_name = 'management/farmer_create.html'
    fields = [
        'farm_name', 'owner_name', 'phone', 'email', 'address',
        'farm_area', 'certification_type', 'is_active'
    ]
    success_url = reverse_lazy('management:farmer_list')


class FarmerDetailView(LoginRequiredMixin, DetailView):
    """Chi tiết trang trại"""
    model = Farmer
    template_name = 'management/farmer_detail.html'
    context_object_name = 'farmer'


class FarmerUpdateView(LoginRequiredMixin, UpdateView):
    """Cập nhật trang trại"""
    model = Farmer
    template_name = 'management/farmer_edit.html'
    fields = [
        'farm_name', 'owner_name', 'phone', 'email', 'address',
        'farm_area', 'certification_type', 'is_active'
    ]
    success_url = reverse_lazy('management:farmer_list')


class FarmerDeleteView(LoginRequiredMixin, DeleteView):
    """Xóa trang trại"""
    model = Farmer
    success_url = reverse_lazy('management:farmer_list')


# Customer Views
class CustomerListView(LoginRequiredMixin, ListView):
    """Danh sách khách hàng"""
    model = CustomerModel
    template_name = 'management/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Customer.objects.all()
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(customer_code__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Filters

        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        return context


class CustomerCreateView(LoginRequiredMixin, CreateView):
    """Tạo khách hàng mới"""
    model = Customer
    template_name = 'management/customer_create.html'
    fields = [
        'customer_code', 'name', 'email', 'phone',
        'address', 'website','representative_name',
        'representative_title', 'priority', '',
        'is_active', 'notes'
    ]
    success_url = reverse_lazy('management:customer_list')
    
    def form_valid(self, form):
        try:
            # Custom validation
            if form.instance.phone == 'business' and not form.instance.tax_code:
                form.add_error('tax_code', 'Mã số thuế là bắt buộc cho doanh nghiệp')
                return self.form_invalid(form)
                
            messages.success(self.request, f'Tạo khách hàng "{form.instance.name}" thành công!')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f'Có lỗi xảy ra: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Có lỗi trong quá trình tạo khách hàng. Vui lòng kiểm tra lại.')
        return super().form_invalid(form)


class CustomerDetailView(LoginRequiredMixin, DetailView):
    """Chi tiết khách hàng"""
    model = Customer
    template_name = 'management/customer_detail.html'
    context_object_name = 'customer'


class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    """Cập nhật khách hàng"""
    model = Customer
    template_name = 'management/customer_edit.html'
    fields = [
        'customer_code', 'name',  'avatar', 'email', 'phone',
        'address', 'website',  'representative_name',
        'representative_title',
        'is_active', 'notes'
    ]
    success_url = reverse_lazy('management:customer_list')


class CustomerDeleteView(LoginRequiredMixin, DeleteView):
    """Xóa khách hàng"""
    model = Customer
    success_url = reverse_lazy('management:customer_list')


# Export/Import Views
def export_data(request, data_type):
    """Xuất dữ liệu ra Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    if data_type == 'employees':
        response['Content-Disposition'] = 'attachment; filename=nhan_vien.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nhân viên"
        
        # Headers
        headers = ['Mã NV', 'Họ tên', 'Email', 'Điện thoại', 'Phòng ban', 'Chức vụ', 'Ngày vào làm', 'Trạng thái']
        ws.append(headers)
        
        # Data
        for emp in Employee.objects.all():
            ws.append([
                emp.employee_id,
                emp.full_name,
                emp.email,
                emp.phone,
                emp.get_department_display(),
                emp.get_position_display(),
                emp.hire_date.strftime('%d/%m/%Y'),
                'Đang làm việc' if emp.is_active else 'Đã nghỉ'
            ])
    
    elif data_type == 'companies':
        response['Content-Disposition'] = 'attachment; filename=cong_ty.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Công ty"
        
        # Headers
        headers = ['Tên trang trại', 'Mã trang trại', 'Chủ trang trại', 'Email', 'Điện thoại', 'Địa chỉ', 'Diện tích', 'Chứng nhận', 'Trạng thái']
        ws.append(headers)
        
        # Data
        for farmer in Farmer.objects.all():
            ws.append([
                farmer.farm_name,
                farmer.farmer_code,
                farmer.owner_name,
                farmer.email or '',
                farmer.phone or '',
                farmer.address or '',
                f"{farmer.farm_area} hecta" if farmer.farm_area else '',
                farmer.get_certification_type_display() if farmer.certification_type else '',
                'Hoạt động' if farmer.is_active else 'Tạm dừng'
            ])
    
    elif data_type == 'customers':
        response['Content-Disposition'] = 'attachment; filename=khach_hang.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Khách hàng"
        
        # Headers
        headers = ['Mã KH', 'Tên khách hàng', 'Loại', 'Email', 'Điện thoại', 'Địa chỉ', 'Ưu tiên', 'VIP', 'Trạng thái']
        ws.append(headers)
        
        # Data
        for customer in Customer.objects.all():
            ws.append([
                customer.customer_code,
                customer.name,
                customer.email or '',
                customer.phone or '',
                customer.address or '',
                customer.get_priority_display(),
                'Hoạt động' if customer.is_active else 'Tạm dừng'
            ])
    
    wb.save(response)
    return response


def import_data(request):
    """Nhập dữ liệu từ file"""
    if request.method == 'POST':
        import_type = request.POST.get('import_type')
        import_file = request.FILES.get('import_file')
        
        if not import_file:
            return JsonResponse({'success': False, 'message': 'Vui lòng chọn file'})
        
        try:
            # Đọc file Excel
            wb = openpyxl.load_workbook(import_file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            
            # Bỏ qua dòng header
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    if import_type == 'employees':
                        # Import nhân viên
                        Employee.objects.create(
                            employee_id=row[0],
                            first_name=row[1].split()[0] if row[1] else '',
                            last_name=' '.join(row[1].split()[1:]) if row[1] and len(row[1].split()) > 1 else row[1],
                            email=row[2],
                            phone=row[3],
                            department='staff',  # Default
                            position='staff',  # Default
                            hire_date=datetime.now().date()
                        )
                    elif import_type == 'customers':
                        # Import khách hàng
                        Customer.objects.create(
                            name=row[1],
                            email=row[3] if row[3] else '',
                            phone=row[4] if row[4] else '',
                            address=row[5] if row[5] else ''
                        )
                    
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    continue
            
            return JsonResponse({
                'success': True,
                'message': f'Nhập thành công {success_count} bản ghi. Lỗi: {error_count} bản ghi.'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Lỗi đọc file: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})
